import os
import json
import csv
from bs4 import BeautifulSoup
import requests
import re
import sys
from openpyxl import Workbook, load_workbook
from datetime import datetime

# 新しいCSVファイルのパス
csv_file_path = "新しいファイル.csv"
current_directory = os.getcwd()
html_relative_path = "load.html"
urlcsv_relative_path = "url.csv"
# HTMLファイルの絶対パスを生成
html_file_path = os.path.join(current_directory, html_relative_path)
item_name = {0: ""}
csv_input_data = []
# CHFの正規表現パターン
chf_pattern = re.compile(r"\bCHF\b")
testcheck_pattern = re.compile(r"\b未使用品\b")
cyukocheck_pattern = re.compile(r"\b中古品\b\s*(?P<description>.+)", re.DOTALL)


brbproducts_list_array = []
brbproducts_list_array_index = 0
brbproduct = ""
single_rowdata = []


# 日本語を取り除く正規表現の関数
def remove_japanese(text):
    pattern = r"[\p{Script=Hiragana}\p{Script=Katakana}\p{Script=Han}]"
    return re.sub(pattern, "", text)


def arraysort_check(array, text):
    misiyouhinncheck = "未使用品"
    cyukohinncheck = "中古品"
    if text in misiyouhinncheck:
        array.append(text)
        return array
    elif text in cyukohinncheck:
        array.append("")
        array.append(text)
        return array
    else:
        array.append(text)
    return array


# 金額を取り除く、正規表現の関数
def extract_amount_or_text(text):
    # テキストが文字列でない場合はそのまま返す。
    if not isinstance(text, str):
        print("→は文字列ではない", text)
        return text
    returntextarrays = []
    alphabet_number_pattern = r"\b(?:\d{2,}[A-Za-z]+|\d{2,})\b"
    alphabet_number_Extract = re.findall(alphabet_number_pattern, text)
    # サイズの正規表現
    size_pattern = r"\b\d{2}(?:mm)?\b"
    size_Extract = re.findall(size_pattern, text)

    # テキストでリファレンスナンバーのターンだと、色などの情報を配列で返す。
    if alphabet_number_Extract:
        print("リファレンスナンバー", text)
        # リファレンスナンバー以外はsubで抽出する
        returntextarrays.append(alphabet_number_Extract[0])
        itemname = re.sub(alphabet_number_pattern, "", text)
        itemname = re.sub(size_pattern, "", itemname)
        returntextarrays.append(itemname)
        # alphabet_number_Extract.append(itemname)
        # サイズの配列も入れる。
    # !この時点で、製品名や金額であり、配列ではない
    else:
        print("リファレンスナンバーでない", text)
        return text

    if size_Extract:
        # テキストからサイズを削除する。
        returntextarrays.append(size_Extract[0])
    else:
        # サイズがなくても空の配列を追加する。
        returntextarrays.append("")
    return returntextarrays
    # ここまで来て配列の長さが0
    # 金額をマッチさせる正規表現パターン
    pattern = r"\$?(\d+(?:,\d+)*)"  # 小数点以下がある場合を考慮する
    amounts = re.findall(pattern, text)

    if amounts:
        print("マッチした元のテキスト→", text)
        print("マッチしてる", amounts)
        return amounts[0]  # 最初にマッチした金額を返す
    else:
        print("マッチしていない", text)
        return text  # 金額が含まれていない場合は元のテキストを返す


# !エクセルデータ設定
# エクセルのヘッダ－データ
data = ["製品名", "リファレンスNO", "最高価格", "最安価格", "ブレスレット", "その他"]
# ファイル名に日付を組み込む
# 現在の日付を取得
today_date = datetime.now().strftime("%Y%m%d")
file_name = f"output_{today_date}.xlsx"
if not os.path.exists(file_name):
    # Excelブックの作成
    wb = Workbook()
    ws = wb.active
    # ヘッダー行を追加
    ws.append(["種類", "リファレンスNO", "製品名", "サイズ", "未使用品", "中古品"])
else:
    # ファイルが存在する場合は既存のファイルを読み込み
    wb = load_workbook(file_name)
    ws = wb.active

# !エクセルデータ設定ここまで
with open(html_file_path, "r", encoding="utf-8") as file:
    soup = BeautifulSoup(file, "html.parser")
    brbproducts_list = soup.find_all(class_="MarketSearch2_itemDiv")  # クラス名を修正
    brbproducts_list = soup.find_all(class_="MarketSearch2_item")  # クラス名を修正

    page_array_index = 0
    index = 0
    for item in brbproducts_list:
        # テキストを走査
        input_text = item.text
        text = input_text.splitlines()
        row_item = []
        index += 1

        text = list(filter(None, text))

        # print(text)
        # print("---------------")
        for item in text:
            # !以下各アイテムのデータ型。
            # ?種類=文字列
            # ?"リファレンスNO", "製品名", "サイズ"=配列
            # ?未使用品および中古品=文字列
            cyuko_match = cyukocheck_pattern.search(item)
            print(f"{len(text)}はアイテムの数です。")
            if cyuko_match:
                print("金額を取得", description=cyuko_match.group("description"))
                item = re.sub(chf_pattern, "", item)
                row_item.insert(4, item)
                continue
            else:
                print("→は金額でない", item)
            #
            if "中古品" in item:
                print("中古品→", item)
                item = item.replace("中古品", "")
                if len(text) == 3:
                    row_item.append("")
                    row_item.append(item)
                else:
                    row_item.append(item)
                continue
            else:
                print("中古品でない→", item)
            if "未使用品" in item:
                print("未使用品", item)
                item = item.replace("未使用品", "")
                row_item.append(item)
                continue
            else:
                print("未使用品でない")

            # !アイテムが返ってくる配列の場合を分岐する。
            item = extract_amount_or_text(item)
            # itemの配列判定して、配列の長さだけrow_itemに追加する。
            #!配列で返ってくる場合、リファレンスナンバー、アイテム情報、サイズが返ってくる想定。
            if isinstance(item, list):
                # row_item.append(item[0])
                # row_item.append(item[1])
                # print("リファレンスナンバー", item[0])
                # print("そのたアイテム情報", item[1])
                for i in item:
                    row_item.append(i)
            else:
                # 配列ではない、金額などの文字列および数値型
                print("リファレンスナンバー、アイテム名サイズ以外→", item)
                # 未使用品は配列に通常n+1番目に入れる。
                if item in "中古品":
                    print("中古品チェック", item)
                    row_item.append("")
                    row_item.append(item)
                    continue
                # 配列をまわして中古品は配列n+2番目に入れる。

                row_item.append(item)
            # !以下でエクセルファイルに入稿する。
            # ws.append(itemmatches)
            #     if not item.strip():
            #         continue
            #     else:
            #         brbproduct += brbproduct + "\n" + item

            # chf_match = testcheck_pattern.search(item)
            # if chf_match:
            #     #             brbproducts_list_array.append(brbproduct)
            #     #             print(brbproducts_list_array[brbproducts_list_array_index])
            #     #             brbproducts_list_array_index += 1
            #     #             brbproduct = ""
            #     print("----------------")
        ws.append(row_item)
        print("-------ここでエクセルファイルに入稿する---------")

wb.save(file_name)
